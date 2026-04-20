"""CLI smoke tests for convert_raghpo_paper_data.py."""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook

SCRIPTS_DIR = Path(__file__).resolve().parents[3] / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import convert_raghpo_paper_data  # noqa: E402


def _write_inputs(tmp_path: Path) -> tuple[Path, Path, Path]:
    csv_path = tmp_path / "Test_Cases.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["Case", "clinical_note"])
        writer.writeheader()
        writer.writerow(
            {
                "Case": "1",
                "clinical_note": "Patient has obesity and seizures since infancy.",
            }
        )

    workbook_path = tmp_path / "RAG-HPO Tests and Data Analysis copy.xlsx"
    workbook = Workbook()
    csc_input = workbook.active
    csc_input.title = "CSC Input"
    csc_input.append(["Case", "clinical_note"])
    csc_input.append([1, "Patient has obesity and seizures since infancy."])
    gsc_input = workbook.create_sheet("GSC Input")
    gsc_input.append(["patient_id", "ID", "clinical_note"])
    gsc_input.append([2, 263442, "Child demonstrates microcephaly on examination."])
    csc_annotations = workbook.create_sheet("CSC Manual Annotations")
    csc_annotations.append(["Patient ID", "hpo_description", "hpo_term"])
    csc_annotations.append([1, "obesity", "HP:0001513"])
    gsc_annotations = workbook.create_sheet("GSC Manual Annotations")
    gsc_annotations.append(
        ["Patient ID", "ID", "hpo_description", "hpo_term", "Category"]
    )
    gsc_annotations.append([2, 263442, "microcephaly", "HP:0000252", "Abnormal"])
    workbook.create_sheet("Premium LLM Output")
    workbook.create_sheet("Premium LLM Comparison")
    workbook.create_sheet("CSC Comparison Results")
    workbook.create_sheet("GSC Comparison Results")
    workbook.save(workbook_path)

    hpo_terms_path = tmp_path / "hpo_terms.tsv"
    hpo_terms_path.write_text(
        "id\tname\nHP:0001513\tObesity\nHP:0000252\tMicrocephaly\n",
        encoding="utf-8",
    )

    return workbook_path, csv_path, hpo_terms_path


def _write_hpo_json(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "graphs": [
                    {
                        "nodes": [
                            {
                                "id": "http://purl.obolibrary.org/obo/HP_0001513",
                                "lbl": "Obesity",
                                "meta": {},
                            },
                            {
                                "id": "http://purl.obolibrary.org/obo/HP_0000252",
                                "lbl": "Microcephaly",
                                "meta": {},
                            },
                        ]
                    }
                ]
            }
        )
        + "\n",
        encoding="utf-8",
    )


def test_cli_converts_selected_dataset(tmp_path: Path, monkeypatch) -> None:
    workbook_path, csv_path, hpo_terms_path = _write_inputs(tmp_path)
    output_dir = tmp_path / "converted"

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "convert_raghpo_paper_data.py",
            "--workbook",
            str(workbook_path),
            "--test-cases-csv",
            str(csv_path),
            "--hpo-terms",
            str(hpo_terms_path),
            "--output",
            str(output_dir),
            "--dataset",
            "CSC",
        ],
    )

    convert_raghpo_paper_data.main()

    output_file = output_dir / "CSC" / "annotations" / "CSC_1.json"
    assert output_file.exists()
    payload = json.loads(output_file.read_text(encoding="utf-8"))
    assert payload["doc_id"] == "CSC_1"


def test_cli_requires_hpo_json_when_normalizing_obsolete_ids(
    tmp_path: Path, monkeypatch
) -> None:
    workbook_path, csv_path, hpo_terms_path = _write_inputs(tmp_path)
    output_dir = tmp_path / "converted"

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "convert_raghpo_paper_data.py",
            "--workbook",
            str(workbook_path),
            "--test-cases-csv",
            str(csv_path),
            "--hpo-terms",
            str(hpo_terms_path),
            "--output",
            str(output_dir),
            "--dataset",
            "CSC",
            "--normalize-obsolete-ids",
        ],
    )

    try:
        convert_raghpo_paper_data.main()
    except SystemExit as exc:
        assert exc.code == 1
    else:
        raise AssertionError("Expected SystemExit when --hpo-json is omitted")


def test_cli_accepts_hpo_json_without_hpo_terms(tmp_path: Path, monkeypatch) -> None:
    workbook_path, csv_path, _ = _write_inputs(tmp_path)
    hpo_json_path = tmp_path / "hp.json"
    _write_hpo_json(hpo_json_path)
    output_dir = tmp_path / "converted"

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "convert_raghpo_paper_data.py",
            "--workbook",
            str(workbook_path),
            "--test-cases-csv",
            str(csv_path),
            "--hpo-json",
            str(hpo_json_path),
            "--output",
            str(output_dir),
            "--dataset",
            "CSC",
        ],
    )

    convert_raghpo_paper_data.main()

    output_file = output_dir / "CSC" / "annotations" / "CSC_1.json"
    assert output_file.exists()
    payload = json.loads(output_file.read_text(encoding="utf-8"))
    assert payload["annotations"][0]["label"] == "Obesity"


def test_cli_requires_input_paths_without_download_mode(
    tmp_path: Path, monkeypatch
) -> None:
    output_dir = tmp_path / "converted"
    hpo_json_path = tmp_path / "hp.json"
    _write_hpo_json(hpo_json_path)

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "convert_raghpo_paper_data.py",
            "--hpo-json",
            str(hpo_json_path),
            "--output",
            str(output_dir),
            "--dataset",
            "CSC",
        ],
    )

    try:
        convert_raghpo_paper_data.main()
    except SystemExit as exc:
        assert exc.code == 1
    else:
        raise AssertionError(
            "Expected SystemExit when workbook/test-cases are omitted without --download-source"
        )


def test_cli_can_download_source_files_when_enabled(
    tmp_path: Path, monkeypatch
) -> None:
    _, _, _ = _write_inputs(tmp_path)
    hpo_json_path = tmp_path / "hp.json"
    _write_hpo_json(hpo_json_path)
    output_dir = tmp_path / "converted"
    download_dir = tmp_path / "downloaded"

    def fake_download(url: str, destination: Path) -> Path:
        destination.parent.mkdir(parents=True, exist_ok=True)
        if destination.suffix == ".csv":
            destination.write_text(
                "Case,clinical_note\n"
                "1,Patient has obesity and seizures since infancy.\n",
                encoding="utf-8",
            )
        else:
            workbook = Workbook()
            csc_input = workbook.active
            csc_input.title = "CSC Input"
            csc_input.append(["Case", "clinical_note"])
            csc_input.append([1, "Patient has obesity and seizures since infancy."])
            gsc_input = workbook.create_sheet("GSC Input")
            gsc_input.append(["patient_id", "ID", "clinical_note"])
            gsc_input.append(
                [2, 263442, "Child demonstrates microcephaly on examination."]
            )
            csc_annotations = workbook.create_sheet("CSC Manual Annotations")
            csc_annotations.append(["Patient ID", "hpo_description", "hpo_term"])
            csc_annotations.append([1, "obesity", "HP:0001513"])
            gsc_annotations = workbook.create_sheet("GSC Manual Annotations")
            gsc_annotations.append(
                ["Patient ID", "ID", "hpo_description", "hpo_term", "Category"]
            )
            gsc_annotations.append(
                [2, 263442, "microcephaly", "HP:0000252", "Abnormal"]
            )
            workbook.create_sheet("Premium LLM Output")
            workbook.create_sheet("Premium LLM Comparison")
            workbook.create_sheet("CSC Comparison Results")
            workbook.create_sheet("GSC Comparison Results")
            workbook.save(destination)
        return destination

    monkeypatch.setattr(convert_raghpo_paper_data, "_download_file", fake_download)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "convert_raghpo_paper_data.py",
            "--download-source",
            "--download-dir",
            str(download_dir),
            "--hpo-json",
            str(hpo_json_path),
            "--output",
            str(output_dir),
            "--dataset",
            "CSC",
        ],
    )

    convert_raghpo_paper_data.main()

    assert (download_dir / convert_raghpo_paper_data.DEFAULT_WORKBOOK_NAME).exists()
    assert (download_dir / convert_raghpo_paper_data.DEFAULT_TEST_CASES_NAME).exists()
    assert (output_dir / "CSC" / "annotations" / "CSC_1.json").exists()
