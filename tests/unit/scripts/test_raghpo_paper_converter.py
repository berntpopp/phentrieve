"""Tests for converting released RAG-HPO paper data into benchmark JSON."""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

SCRIPTS_DIR = Path(__file__).resolve().parents[3] / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from raghpo_paper_converter import RagHpoPaperConverter  # noqa: E402


def _write_hpo_terms(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "id\tname",
                "HP:0001513\tObesity",
                "HP:0001250\tSeizure",
                "HP:0000252\tMicrocephaly",
                "HP:0001249\tIntellectual disability",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def _write_hpo_json(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "graphs": [
                    {
                        "nodes": [
                            {
                                "id": "http://purl.obolibrary.org/obo/HP_0001513",
                                "lbl": "Obesity",
                                "meta": {},
                            },
                            {
                                "id": "http://purl.obolibrary.org/obo/HP_0001250",
                                "lbl": "Seizure",
                                "meta": {},
                            },
                            {
                                "id": "http://purl.obolibrary.org/obo/HP_0002355",
                                "lbl": "obsolete Difficulty walking",
                                "meta": {
                                    "deprecated": True,
                                    "basicPropertyValues": [
                                        {
                                            "pred": "http://purl.obolibrary.org/obo/IAO_0100001",
                                            "val": "HP:0001288",
                                        }
                                    ],
                                },
                            },
                            {
                                "id": "http://purl.obolibrary.org/obo/HP_0001288",
                                "lbl": "Gait disturbance",
                                "meta": {},
                            },
                            {
                                "id": "http://purl.obolibrary.org/obo/HP_0001263",
                                "lbl": "obsolete Developmental delay",
                                "meta": {
                                    "deprecated": True,
                                    "basicPropertyValues": [
                                        {
                                            "pred": "http://purl.obolibrary.org/obo/IAO_0100001",
                                            "val": "HP:0025356",
                                        }
                                    ],
                                },
                            },
                            {
                                "id": "http://purl.obolibrary.org/obo/HP_0025356",
                                "lbl": "obsolete Psychomotor retardation",
                                "meta": {
                                    "deprecated": True,
                                    "basicPropertyValues": [
                                        {
                                            "pred": "http://purl.obolibrary.org/obo/IAO_0100001",
                                            "val": "HP:0001249",
                                        }
                                    ],
                                },
                            },
                            {
                                "id": "http://purl.obolibrary.org/obo/HP_0001249",
                                "lbl": "Intellectual disability",
                                "meta": {},
                            },
                            {
                                "id": "http://purl.obolibrary.org/obo/HP_1234567",
                                "lbl": "obsolete Synthetic no-replacement phenotype",
                                "meta": {"deprecated": True},
                            },
                        ]
                    }
                ]
            }
        )
        + "\n",
        encoding="utf-8",
    )


def _write_test_cases_csv(path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["Case", "clinical_note"])
        writer.writeheader()
        writer.writerow(
            {
                "Case": "1",
                "clinical_note": "Patient has obesity and seizures since infancy.",
            }
        )


def _write_test_cases_csv_with_unannotated_case(path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["Case", "clinical_note"])
        writer.writeheader()
        writer.writerow(
            {
                "Case": "1",
                "clinical_note": "Patient has obesity and seizures since infancy.",
            }
        )
        writer.writerow(
            {
                "Case": "2",
                "clinical_note": "Unannotated control case with no manual labels.",
            }
        )


def _write_test_cases_csv_with_bom(path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["Case", "clinical_note"])
        writer.writeheader()
        writer.writerow(
            {
                "Case": "1",
                "clinical_note": "Patient has obesity and seizures since infancy.",
            }
        )


def _write_raghpo_workbook(path: Path, unmatched_phrase: str = "microcephaly") -> None:
    workbook = Workbook()

    csc_input = workbook.active
    csc_input.title = "CSC Input"
    csc_input.append(["Case", "clinical_note"])
    csc_input.append([1, "Patient has obesity and seizures since infancy."])

    gsc_input = workbook.create_sheet("GSC Input")
    gsc_input.append(["patient_id", "ID", "clinical_note"])
    gsc_input.append([2, 263442, "Child demonstrates microcephaly on examination."])

    csc_annotations = workbook.create_sheet("CSC Manual Annotations")
    csc_annotations.append(["Patient ID", "hpo_description", "hpo_term"])
    csc_annotations.append([1, "obesity", "HP:0001513"])
    csc_annotations.append([1, "seizures", "HP:0001250"])

    gsc_annotations = workbook.create_sheet("GSC Manual Annotations")
    gsc_annotations.append(
        ["Patient ID", "ID", "hpo_description", "hpo_term", "Category"]
    )
    gsc_annotations.append([2, 263442, unmatched_phrase, "HP:0000252", "Abnormal"])

    workbook.create_sheet("Premium LLM Output")
    workbook.create_sheet("Premium LLM Comparison")
    workbook.create_sheet("CSC Comparison Results")
    workbook.create_sheet("GSC Comparison Results")

    workbook.save(path)


def _write_raghpo_workbook_with_joined_hpo_ids(path: Path) -> None:
    workbook = Workbook()

    csc_input = workbook.active
    csc_input.title = "CSC Input"
    csc_input.append(["Case", "clinical_note"])
    csc_input.append([1, "Patient has obesity and seizures since infancy."])

    gsc_input = workbook.create_sheet("GSC Input")
    gsc_input.append(["patient_id", "ID", "clinical_note"])

    csc_annotations = workbook.create_sheet("CSC Manual Annotations")
    csc_annotations.append(["Patient ID", "hpo_description", "hpo_term"])
    csc_annotations.append([1, "obesity", "HP:0001513, HP:0001250"])

    gsc_annotations = workbook.create_sheet("GSC Manual Annotations")
    gsc_annotations.append(
        ["Patient ID", "ID", "hpo_description", "hpo_term", "Category"]
    )

    workbook.create_sheet("Premium LLM Output")
    workbook.create_sheet("Premium LLM Comparison")
    workbook.create_sheet("CSC Comparison Results")
    workbook.create_sheet("GSC Comparison Results")

    workbook.save(path)


@pytest.fixture
def raghpo_source_files(tmp_path: Path) -> dict[str, Path]:
    csv_path = tmp_path / "Test_Cases.csv"
    workbook_path = tmp_path / "RAG-HPO Tests and Data Analysis copy.xlsx"
    hpo_terms_path = tmp_path / "hpo_terms.tsv"

    _write_test_cases_csv(csv_path)
    _write_raghpo_workbook(workbook_path)
    _write_hpo_terms(hpo_terms_path)

    return {
        "csv": csv_path,
        "workbook": workbook_path,
        "hpo_terms": hpo_terms_path,
    }


def test_converter_writes_csc_documents_in_existing_benchmark_shape(
    tmp_path: Path, raghpo_source_files: dict[str, Path]
) -> None:
    output_dir = tmp_path / "converted"
    converter = RagHpoPaperConverter(hpo_terms_path=raghpo_source_files["hpo_terms"])

    report = converter.convert(
        workbook_path=raghpo_source_files["workbook"],
        test_cases_csv_path=raghpo_source_files["csv"],
        output_root=output_dir,
        dataset="CSC",
    )

    output_file = output_dir / "CSC" / "annotations" / "CSC_1.json"
    assert output_file.exists()

    converted = json.loads(output_file.read_text(encoding="utf-8"))
    assert converted["doc_id"] == "CSC_1"
    assert converted["language"] == "en"
    assert converted["source"] == "rag_hpo_paper"
    assert converted["metadata"]["dataset"] == "CSC"
    assert converted["metadata"]["original_doc_id"] == "1"
    assert converted["metadata"]["num_annotations"] == 2
    assert converted["full_text"] == "Patient has obesity and seizures since infancy."
    assert [item["hpo_id"] for item in converted["annotations"]] == [
        "HP:0001513",
        "HP:0001250",
    ]
    assert converted["annotations"][0]["label"] == "Obesity"
    assert converted["annotations"][0]["assertion_status"] == "affirmed"
    assert converted["annotations"][0]["evidence_spans"] == [
        {
            "start_char": 12,
            "end_char": 19,
            "text_snippet": "obesity",
        }
    ]
    assert report["summary"]["total_documents"] == 1
    assert report["datasets"]["CSC"]["documents"] == 1


def test_converter_can_resolve_labels_from_hpo_json_without_terms_tsv(
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "Test_Cases.csv"
    workbook_path = tmp_path / "RAG-HPO Tests and Data Analysis copy.xlsx"
    hpo_json_path = tmp_path / "hp.json"
    output_dir = tmp_path / "converted"

    _write_test_cases_csv(csv_path)
    _write_raghpo_workbook(workbook_path)
    _write_hpo_json(hpo_json_path)

    converter = RagHpoPaperConverter(hpo_terms_path=None, hpo_json_path=hpo_json_path)
    converter.convert(
        workbook_path=workbook_path,
        test_cases_csv_path=csv_path,
        output_root=output_dir,
        dataset="CSC",
    )

    output_file = output_dir / "CSC" / "annotations" / "CSC_1.json"
    converted = json.loads(output_file.read_text(encoding="utf-8"))
    assert converted["annotations"][0]["label"] == "Obesity"


def test_converter_writes_gsc_documents_from_workbook_inputs(
    tmp_path: Path, raghpo_source_files: dict[str, Path]
) -> None:
    output_dir = tmp_path / "converted"
    converter = RagHpoPaperConverter(hpo_terms_path=raghpo_source_files["hpo_terms"])

    converter.convert(
        workbook_path=raghpo_source_files["workbook"],
        test_cases_csv_path=raghpo_source_files["csv"],
        output_root=output_dir,
        dataset="GSC",
    )

    output_file = output_dir / "GSC" / "annotations" / "GSC_2.json"
    converted = json.loads(output_file.read_text(encoding="utf-8"))

    assert converted["doc_id"] == "GSC_2"
    assert converted["metadata"]["dataset"] == "GSC"
    assert converted["metadata"]["original_doc_id"] == "263442"
    assert converted["annotations"] == [
        {
            "hpo_id": "HP:0000252",
            "label": "Microcephaly",
            "assertion_status": "affirmed",
            "evidence_spans": [
                {
                    "start_char": 19,
                    "end_char": 31,
                    "text_snippet": "microcephaly",
                }
            ],
        }
    ]


def test_converter_keeps_unmatched_evidence_as_empty_span_list_and_reports_it(
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "Test_Cases.csv"
    workbook_path = tmp_path / "RAG-HPO Tests and Data Analysis copy.xlsx"
    hpo_terms_path = tmp_path / "hpo_terms.tsv"
    output_dir = tmp_path / "converted"

    _write_test_cases_csv(csv_path)
    _write_raghpo_workbook(workbook_path, unmatched_phrase="cephalic narrowing")
    _write_hpo_terms(hpo_terms_path)

    converter = RagHpoPaperConverter(hpo_terms_path=hpo_terms_path)
    report = converter.convert(
        workbook_path=workbook_path,
        test_cases_csv_path=csv_path,
        output_root=output_dir,
        dataset="GSC",
    )

    output_file = output_dir / "GSC" / "annotations" / "GSC_2.json"
    converted = json.loads(output_file.read_text(encoding="utf-8"))

    assert converted["annotations"][0]["evidence_spans"] == []
    assert report["summary"]["warnings"] == 1
    assert "cephalic narrowing" in report["warnings"][0]


def test_converter_reads_upstream_utf8_bom_test_cases_csv(tmp_path: Path) -> None:
    csv_path = tmp_path / "Test_Cases.csv"
    workbook_path = tmp_path / "RAG-HPO Tests and Data Analysis copy.xlsx"
    hpo_terms_path = tmp_path / "hpo_terms.tsv"
    output_dir = tmp_path / "converted"

    _write_test_cases_csv_with_bom(csv_path)
    _write_raghpo_workbook(workbook_path)
    _write_hpo_terms(hpo_terms_path)

    converter = RagHpoPaperConverter(hpo_terms_path=hpo_terms_path)
    report = converter.convert(
        workbook_path=workbook_path,
        test_cases_csv_path=csv_path,
        output_root=output_dir,
        dataset="CSC",
    )

    assert report["summary"]["total_documents"] == 1
    assert (output_dir / "CSC" / "annotations" / "CSC_1.json").exists()


def test_converter_filters_csc_to_annotated_cases_only(tmp_path: Path) -> None:
    csv_path = tmp_path / "Test_Cases.csv"
    workbook_path = tmp_path / "RAG-HPO Tests and Data Analysis copy.xlsx"
    hpo_terms_path = tmp_path / "hpo_terms.tsv"
    output_dir = tmp_path / "converted"

    _write_test_cases_csv_with_unannotated_case(csv_path)
    _write_raghpo_workbook(workbook_path)
    _write_hpo_terms(hpo_terms_path)

    converter = RagHpoPaperConverter(hpo_terms_path=hpo_terms_path)
    report = converter.convert(
        workbook_path=workbook_path,
        test_cases_csv_path=csv_path,
        output_root=output_dir,
        dataset="CSC",
    )

    csc_annotations_dir = output_dir / "CSC" / "annotations"
    assert (csc_annotations_dir / "CSC_1.json").exists()
    assert not (csc_annotations_dir / "CSC_2.json").exists()
    assert report["summary"]["total_documents"] == 1
    assert report["datasets"]["CSC"]["documents"] == 1


def test_converter_splits_comma_joined_hpo_ids_into_separate_annotations(
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "Test_Cases.csv"
    workbook_path = tmp_path / "RAG-HPO Tests and Data Analysis copy.xlsx"
    hpo_terms_path = tmp_path / "hpo_terms.tsv"
    output_dir = tmp_path / "converted"

    _write_test_cases_csv(csv_path)
    _write_raghpo_workbook_with_joined_hpo_ids(workbook_path)
    _write_hpo_terms(hpo_terms_path)

    converter = RagHpoPaperConverter(hpo_terms_path=hpo_terms_path)
    report = converter.convert(
        workbook_path=workbook_path,
        test_cases_csv_path=csv_path,
        output_root=output_dir,
        dataset="CSC",
    )

    converted = json.loads(
        (output_dir / "CSC" / "annotations" / "CSC_1.json").read_text(encoding="utf-8")
    )
    assert [item["hpo_id"] for item in converted["annotations"]] == [
        "HP:0001513",
        "HP:0001250",
    ]
    assert report["summary"]["total_annotations"] == 2


def test_converter_can_normalize_obsolete_hpo_ids_when_enabled(tmp_path: Path) -> None:
    csv_path = tmp_path / "Test_Cases.csv"
    workbook_path = tmp_path / "RAG-HPO Tests and Data Analysis copy.xlsx"
    hpo_terms_path = tmp_path / "hpo_terms.tsv"
    hpo_json_path = tmp_path / "hp.json"
    output_dir = tmp_path / "converted"

    _write_test_cases_csv(csv_path)
    _write_raghpo_workbook(workbook_path)
    _write_hpo_terms(hpo_terms_path)
    _write_hpo_json(hpo_json_path)

    workbook = Workbook()
    csc_input = workbook.active
    csc_input.title = "CSC Input"
    csc_input.append(["Case", "clinical_note"])
    csc_input.append([1, "Patient has difficulty in walking since infancy."])
    gsc_input = workbook.create_sheet("GSC Input")
    gsc_input.append(["patient_id", "ID", "clinical_note"])
    csc_annotations = workbook.create_sheet("CSC Manual Annotations")
    csc_annotations.append(["Patient ID", "hpo_description", "hpo_term"])
    csc_annotations.append([1, "difficulty in walking", "HP:0002355"])
    gsc_annotations = workbook.create_sheet("GSC Manual Annotations")
    gsc_annotations.append(
        ["Patient ID", "ID", "hpo_description", "hpo_term", "Category"]
    )
    workbook.create_sheet("Premium LLM Output")
    workbook.create_sheet("Premium LLM Comparison")
    workbook.create_sheet("CSC Comparison Results")
    workbook.create_sheet("GSC Comparison Results")
    workbook.save(workbook_path)

    converter = RagHpoPaperConverter(
        hpo_terms_path=hpo_terms_path,
        hpo_json_path=hpo_json_path,
        normalize_obsolete_ids=True,
    )
    report = converter.convert(
        workbook_path=workbook_path,
        test_cases_csv_path=csv_path,
        output_root=output_dir,
        dataset="CSC",
    )

    converted = json.loads(
        (output_dir / "CSC" / "annotations" / "CSC_1.json").read_text(encoding="utf-8")
    )
    assert converted["annotations"][0]["hpo_id"] == "HP:0001288"
    assert converted["annotations"][0]["label"] == "Gait disturbance"
    assert any("HP:0002355 -> HP:0001288" in warning for warning in report["warnings"])


def test_converter_resolves_obsolete_hpo_replacements_transitively(
    tmp_path: Path,
) -> None:
    hpo_terms_path = tmp_path / "hpo_terms.tsv"
    hpo_json_path = tmp_path / "hp.json"
    _write_hpo_terms(hpo_terms_path)
    _write_hpo_json(hpo_json_path)

    lookup = RagHpoPaperConverter(
        hpo_terms_path=hpo_terms_path,
        hpo_json_path=hpo_json_path,
        normalize_obsolete_ids=True,
    )._lookup

    assert lookup.replacement_for("HP:0001263") == "HP:0001249"


def test_converter_can_drop_obsolete_ids_without_replacement_when_enabled(
    tmp_path: Path,
) -> None:
    csv_path = tmp_path / "Test_Cases.csv"
    workbook_path = tmp_path / "RAG-HPO Tests and Data Analysis copy.xlsx"
    hpo_terms_path = tmp_path / "hpo_terms.tsv"
    hpo_json_path = tmp_path / "hp.json"
    output_dir = tmp_path / "converted"

    _write_test_cases_csv(csv_path)
    _write_hpo_terms(hpo_terms_path)
    _write_hpo_json(hpo_json_path)

    workbook = Workbook()
    csc_input = workbook.active
    csc_input.title = "CSC Input"
    csc_input.append(["Case", "clinical_note"])
    csc_input.append([1, "Patient has difficulty in walking since infancy."])
    gsc_input = workbook.create_sheet("GSC Input")
    gsc_input.append(["patient_id", "ID", "clinical_note"])
    csc_annotations = workbook.create_sheet("CSC Manual Annotations")
    csc_annotations.append(["Patient ID", "hpo_description", "hpo_term"])
    csc_annotations.append([1, "difficulty in walking", "HP:1234567"])
    gsc_annotations = workbook.create_sheet("GSC Manual Annotations")
    gsc_annotations.append(
        ["Patient ID", "ID", "hpo_description", "hpo_term", "Category"]
    )
    workbook.create_sheet("Premium LLM Output")
    workbook.create_sheet("Premium LLM Comparison")
    workbook.create_sheet("CSC Comparison Results")
    workbook.create_sheet("GSC Comparison Results")
    workbook.save(workbook_path)

    converter = RagHpoPaperConverter(
        hpo_terms_path=hpo_terms_path,
        hpo_json_path=hpo_json_path,
        normalize_obsolete_ids=True,
        drop_obsolete_without_replacement=True,
    )
    report = converter.convert(
        workbook_path=workbook_path,
        test_cases_csv_path=csv_path,
        output_root=output_dir,
        dataset="CSC",
    )

    converted = json.loads(
        (output_dir / "CSC" / "annotations" / "CSC_1.json").read_text(encoding="utf-8")
    )
    assert converted["annotations"] == []
    assert any(
        "dropped obsolete HPO term HP:1234567" in warning
        for warning in report["warnings"]
    )


def test_converter_closes_excel_workbook_via_context_manager(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook_path = tmp_path / "benchmark.xlsx"
    csv_path = tmp_path / "cases.csv"
    output_dir = tmp_path / "converted"
    hpo_terms_path = tmp_path / "hpo_terms.tsv"
    _write_test_cases_csv(csv_path)
    _write_hpo_terms(hpo_terms_path)

    class FakeExcelFile:
        entered = False
        exited = False

        def __init__(self, path: Path) -> None:
            assert path == workbook_path

        def __enter__(self) -> FakeExcelFile:
            type(self).entered = True
            return self

        def __exit__(self, exc_type, exc, tb) -> None:
            type(self).exited = True

    monkeypatch.setattr("raghpo_paper_converter.pd.ExcelFile", FakeExcelFile)

    converter = RagHpoPaperConverter(hpo_terms_path=hpo_terms_path)
    monkeypatch.setattr(
        converter, "_load_csc_cases", lambda _: {"1": {"full_text": ""}}
    )
    monkeypatch.setattr(converter, "_load_csc_annotations", lambda _: {"1": []})
    monkeypatch.setattr(
        converter,
        "_filter_cases_to_annotations",
        lambda **kwargs: kwargs["cases"],
    )
    monkeypatch.setattr(converter, "_load_gsc_cases", lambda _: {})
    monkeypatch.setattr(converter, "_load_gsc_annotations", lambda _: {})
    monkeypatch.setattr(converter, "_build_documents", lambda **kwargs: [])
    monkeypatch.setattr(converter, "_write_documents", lambda *args, **kwargs: None)

    converter.convert(
        workbook_path=workbook_path,
        test_cases_csv_path=csv_path,
        output_root=output_dir,
        dataset="CSC",
    )

    assert FakeExcelFile.entered is True
    assert FakeExcelFile.exited is True
